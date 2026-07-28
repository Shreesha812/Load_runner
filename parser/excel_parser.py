import logging

from openpyxl import load_workbook

from parser.excel_test_definition_row import ExcelTestDefinitionRow
from parser.workbook_model import WorkbookModel

logger = logging.getLogger(__name__)

# Canonical column names as they appear in the Excel header row.
# Values are case-insensitive and leading/trailing spaces are stripped.
COLUMN_MAP = {
    # URL
    "url": "url",
    # HTTP method
    "method": "method",
    "http method": "method",
    # Request headers
    "headers": "headers",
    # Request body template
    "request template": "request_template",
    "request message template": "request_template",
    # Strategy: sequential / random  (may also carry order info in legacy sheets)
    "variable strategy": "variable_strategy",
    "variable list order": "variable_strategy",   # legacy: "random/sequential" value
    "variable order": "variable_order",
    # Variable definitions with multiple values  e.g. "var1:{a,b,c}, var2:x"
    "list of values": "variable_values",
    "variable values": "variable_values",
    # Single fixed-value variables  e.g. "var1:abc,var2:123"
    "variables": "fixed_variables",
    # Concurrency
    "total concurrent request": "concurrency",
    "concurrency": "concurrency",
    # Response structure assertion
    "response structure": "response_structure",
    # Enable/disable
    "enabled": "enabled",
    "enable": "enabled",
}


class ExcelParser:

    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse(self) -> WorkbookModel:
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        raw_headers = [cell.value for cell in sheet[1]]

        # Build index → field-name mapping from the header row
        index_to_field: dict[int, str] = {}
        for idx, header in enumerate(raw_headers):
            if header is None:
                continue
            normalised = str(header).strip().lower()
            field = COLUMN_MAP.get(normalised)
            if field:
                index_to_field[idx] = field
            else:
                logger.warning("Unrecognised column header '%s' at index %d — ignored.", header, idx)

        # Verify truly required fields are present (optional fields excluded)
        OPTIONAL_FIELDS = {"variable_order", "fixed_variables", "response_structure", "headers"}
        found_fields = set(index_to_field.values())
        required_fields = set(COLUMN_MAP.values()) - OPTIONAL_FIELDS
        missing = required_fields - found_fields
        if missing:
            logger.warning("Excel file is missing required column(s): %s", missing)

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(value is None for value in row):
                continue

            # Build a dict of field → value using the header mapping
            row_data: dict[str, object] = {field: None for field in required_fields}
            for idx, field in index_to_field.items():
                if idx < len(row):
                    row_data[field] = row[idx]

            rows.append(
                ExcelTestDefinitionRow(
                    url=row_data.get("url"),
                    method=row_data.get("method"),
                    headers=row_data.get("headers"),
                    request_template=row_data.get("request_template"),
                    variable_strategy=row_data.get("variable_strategy"),
                    variable_order=row_data.get("variable_order"),
                    variable_values=row_data.get("variable_values"),
                    fixed_variables=row_data.get("fixed_variables"),
                    concurrency=row_data.get("concurrency"),
                    response_structure=row_data.get("response_structure"),
                    enabled=row_data.get("enabled"),
                )
            )

        logger.info(
            "Parsed sheet '%s': %d header column(s), %d data row(s).",
            sheet.title, len(raw_headers), len(rows),
        )

        return WorkbookModel(
            sheet_name=sheet.title,
            headers=raw_headers,
            rows=rows,
        )
